import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from constants import (FINANCIAL_COLUMNS_A,
                       FINANCIAL_COLUMNS_B,
                       HEADER_ROW_NUMBER,
                       KBK_COLUMNS,
                       TABLE_DATA_PATH,
                       TABLE_NEW_PATH)


logging.basicConfig(level=logging.INFO)


def concatenate_kbk_str(data: dict) -> str:
    """Формирует строку с КБК для информации об обновлении."""
    kbk_parts = [
        str(data[kbk_field]) 
        for kbk_field in KBK_COLUMNS 
        if kbk_field in data and data[kbk_field] is not None
    ]
    return f'Изменен КБК: {".".join(kbk_parts)}'


def update_table_a_with_b(table_a_path, table_b_path):
    """
    Обновляет таблицу A данными из таблицы B с сохранением форматирования
    """
    # 1. Загрузка таблиц
    try:
        df_a = pd.read_excel(table_a_path, skiprows=1)
        df_b = pd.read_excel(table_b_path, skiprows=2)
        logging.info("Успешное чтение файлов")
    except Exception as e:
        logging.critical(f'Произошла ошибка при чтении файлов: {e}')
        return
    
    # 2. Заменяем NaN на None для корректной обработки
    df_a = df_a.where(pd.notna(df_a), None)
    df_b = df_b.where(pd.notna(df_b), None)
    
    # 2. Проверяем наиличие всех колонок в обеих таблицах
    # logging.info("Колонки в таблице A:", df_a.columns.tolist())
    # logging.info("Колонки в таблице B:", df_b.columns.tolist())
    
    missing_in_a = [col for col in KBK_COLUMNS if col not in df_a.columns]
    missing_in_b = [col for col in KBK_COLUMNS if col not in df_b.columns]
    
    if missing_in_a:
        logging.critical(f"В таблице A отсутствуют колонки: {missing_in_a}")
        return
    if missing_in_b:
        logging.critical(f"В таблице B отсутствуют колонки: {missing_in_b}")
        return
                
    # 3. Создание словаря для сопоставления колонок
    column_mapping = dict(zip(FINANCIAL_COLUMNS_B, FINANCIAL_COLUMNS_A))
    
    # 4. Объединение таблиц по ключевым полям КБК
    try: 
        merged = pd.merge(
            df_a, 
            df_b, 
            on=KBK_COLUMNS, 
            how='left', 
            suffixes=('_a', '_b')
        )
    except Exception as e:
        logging.critical(f'Произошла ошибка при слиянии таблиц: {e}')
        return
    
    # 5. Загружаем исходный файл с форматированием
    try:
        workbook = load_workbook(table_a_path)
        sheet = workbook.active
    except Exception as e:
        logging.error("Произошла ошибка про загрузке таблицы в openpyxl: {e}")
    
    # 6. Обновление данных с сохранением форматирования
    total_changes = 0
    
    # Находим индексы колонок для обновления
    header_row = HEADER_ROW_NUMBER
    column_indices = {}
    
    # Получаем индексы колонок которые нужно обновить (из таблицы А)
    for idx, cell in enumerate(sheet[header_row], 1):
        if cell.value in FINANCIAL_COLUMNS_A:
            column_indices[cell.value] = idx
    
    # Обновляем данные
    for col_b, col_a in column_mapping.items():
        if col_a not in column_indices:
            continue
        
        # Получаем индекс колонки, которую будем проверять
        col_idx = column_indices[col_a]
        changes_count = 0
        changes_kbk = []
        
        for row_idx, (_, row_data) in enumerate(merged.iterrows(), header_row + 1):
            original_val = sheet.cell(row=row_idx, column=col_idx).value
            new_val = row_data[col_b]
            
            # Проверяем, нужно ли обновлять значение
            if pd.notna(new_val) and original_val != new_val:
                sheet.cell(row=row_idx, column=col_idx).value = new_val
                changes_kbk.append(concatenate_kbk_str(row_data))
                changes_count += 1
                
        
        if changes_count > 0:
            print(f"Обновлено {changes_count} значений в столбце {col_a}")
            for kbk in changes_kbk:
                print(kbk)
            changes_kbk = []
            total_changes += changes_count
    
    # 7. Сохраняем файл с исходным форматированием
    workbook.save(table_a_path)
    print(f"Таблица A обновлена. Всего изменений: {total_changes}.")
    return df_a


# Вызов функции
if __name__ == "__main__":
    updated_df = update_table_a_with_b(
        table_a_path=TABLE_DATA_PATH,
        table_b_path=TABLE_NEW_PATH
    )
