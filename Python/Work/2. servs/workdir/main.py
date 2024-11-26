import requests

from openpyxl import load_workbook
from num2words import num2words

wb = load_workbook('workdir/calc.xlsx')
page = wb["Page_1"]
data = {}


# Функция вывода словаря на экран
def print_data(data):
    for i, j in data.items():
        print(i, j)


# Получаем прописные значения и записываем их в словарь
def get_str_numbers(data):
    for j in data.values():
        j.append(num2words(j[0], lang='ru'))
        j.append(num2words(int(j[1]), lang='ru'))


# Собираем данные из файла exal
for i in range(2, 66):
    key = str(page.cell(row=i, column=1).value)
    data[key] = [
        page.cell(row=i, column=4).value,
        round(page.cell(row=i, column=6).value, 2)
    ]

get_str_numbers(data)
print_data(data)